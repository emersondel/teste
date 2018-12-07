from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import os.path
import openpyxl
from openpyxl import load_workbook
import time
from datetime import date

browser = ""
book = ""
sheet = ""
wb = ""

def openChrome():
    global browser
    #seta as configurações do Google Chrome
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument('--disable-gpu')
    chrome_driver_path = "C:\Python371\Scripts\chromedriver.exe"

    # abre o Chrome para realizar os testes
    browser = webdriver.Chrome(chrome_driver_path, chrome_options=chrome_options)
    #browser = webdriver.Chrome()

    #browser.set_window_size(2000, 2000)

def pesquisaNomeExame(nome_exame):
    global sheet
    for v in range(sheet.max_row):
        if sheet.cell(v+1, 2).value == nome_exame:
                return True
    return False

def insereDadosXLS(nome_exame, prazo, outros_nomes):
    global sheet
    global book
    #copia os dados para uma nova planilha
    #book_new = copy(book)
    #sheet_new = book_new.sheet_by_name("Fleury")

    ultima_linha = sheet.max_row+1

    #insere os dados de nome exame, prazo e outros nomes
    sheet.cell(ultima_linha, 2).value = nome_exame
    sheet.cell(ultima_linha, 4).value = prazo
    sheet.cell(ultima_linha, 6).value = outros_nomes
    sheet.cell(ultima_linha, 7).value = date.today()

    #salva a planilha
    book.save('C:\Projetos\Dasa\Tabela_Exames.xlsx')

def checaPlanilha():
    global book
    global sheet
    global wb
    #verifica se a planilha existe na pasta
    if os.path.exists('C:\Projetos\Dasa\Tabela_Exames.xlsx'):
        #abre a planilha
        book = load_workbook("C:\Projetos\Dasa\Tabela_Exames.xlsx")
        sheet = book.get_sheet_by_name("Fleury")
        #book.sheet_state = 'visible'
    else:
        #Cria a planilha e monta o cabeçalho padrão
        book = openpyxl.Workbook()
        sheet = book.create_sheet('Fleury')

        #deleta as demais abas
        for a in book.worksheets:
            if a.title != 'Fleury':
                book.remove_sheet(a)

        # Títulos das colunas
        titles = ['Categoria', 'Exames', 'Metodologia', 'Prazo', 'Preço', 'Sinonimia', 'Data Inclusão Exame']

        # Escrevendo títulos na primeira linha do arquivo
        for i in range(len(titles)):
            sheet.cell(1, i+1).value = titles[i]

        # Salva a planilha
        book.save('C:\Projetos\Dasa\Tabela_Exames.xlsx')

        # abre a planilha
        book = load_workbook("C:\Projetos\Dasa\Tabela_Exames.xlsx")
        sheet = book.get_sheet_by_name("Fleury")

def coletaDadosExame():
    global browser
    #seta as opções de busca
    tipos = ['A','B',',''C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

    #pesquisa os exames pela inicial
    for i in range(len(tipos)):
        # Navega para a pagina dos exames
        browser.get('http://www.fleury.com.br/exames-e-servicos/medicina-diagnostica/exames-oferecidos/Pages/default.aspx?src_a=0&src_d=10000&BUSCA=' + tipos[i] + '&Tipo=0')
        element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "lista-exames")))

        #diminui o zoom da pagina
        #browser.execute_script("document.body.style.zoom='75%'")

        #coleta o total de exames exibidos na tela
        for y in range(len(browser.find_elements_by_tag_name("article"))):
            #coleta o nome do exame
            nome_exame = browser.find_elements_by_tag_name("article")[y].text.split("\n")[0]
            time.sleep(1)
            #verifica se o exame já foi pesquisado
            if pesquisaNomeExame(nome_exame) == False:
                #clica no exame para coletar os dados necessários
                #browser.find_elements_by_tag_name("article")[y].find_element_by_xpath("//*[@title='Orientações']").click()
                #browser.find_elements_by_xpath("//*[@title='Orientações']")[y].click();
                browser.find_elements_by_xpath("//*[@title='Orientações']")[y].send_keys(Keys.ENTER)
                time.sleep(5)

                # seta para o frame
                browser.switch_to_frame("Iframe1")

                #coleta o prazo
                prazo = browser.find_element_by_id("pPRazoEntrega").text

                #coleta os outros nomes
                outros_nomes = browser.find_element_by_id("dvOutrosNomes").text.replace( "Outros nomes:\n", "" )

                #insere as informações na planilha
                insereDadosXLS(nome_exame, prazo, outros_nomes)

                #tira um print da tela
                browser.get_screenshot_as_file("C:\Projetos\Dasa\Prints_exames\Fleury_" + nome_exame + ".png")

                # Navega para a pagina dos exames
                browser.get(
                    'http://www.fleury.com.br/exames-e-servicos/medicina-diagnostica/exames-oferecidos/Pages/default.aspx?src_a=0&src_d=10000&BUSCA=' +
                    tipos[i] + '&Tipo=0')
                element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, "article")))


#Verifica a planilha
checaPlanilha()

#abre o Chrome
openChrome()

#pesquisa pelos exames no site
coletaDadosExame()

# encerra o browser
browser.quit()
